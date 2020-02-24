import openpyxl
from openpyxl.chart import BarChart,Reference
from xlsxwriter import Workbook
import xlrd
from bs4 import BeautifulSoup
import requests
import selenium
from selenium import webdriver
import pandas as pd
from fake_useragent import UserAgent
from time import sleep
from selenium.common.exceptions import NoSuchElementException

base_url = 'https://www.flipkart.com'
driver = webdriver.Chrome(r'C:\Users\User\Downloads\chromedriver_win32\chromedriver')

def login():
    driver.get('https://www.flipkart.com/account/login')
    usr_name_input = driver.find_element_by_xpath(
        "//div[@id='container']//form/div[@class='_39M2dM JB4AMj']/input[@class='_2zrpKA _1dBPDZ']")
    usr_name_input.send_keys('USERNAME')
    pwd_input = driver.find_element_by_xpath(
        "//div[@id='container']//form/div[@class='_39M2dM JB4AMj']/input[@class='_2zrpKA _3v41xv _1dBPDZ']")
    pwd_input.send_keys('PASSWORD')
    pwd_input.submit()

login()
soup = BeautifulSoup(driver.page_source, 'lxml')
category = [li.a['href'] for li in soup.find_all('li', class_='_1KCOnI _3ZgIXy')]
sub_cat = [base_url + links for links in category]

driver.get(sub_cat[0])
inner_soup = BeautifulSoup(driver.page_source, 'lxml')
product = [a['href'] for a in inner_soup.find_all('a', href=True, attrs={'class': '_31qSD5'})]
final_pg = [base_url + links for links in product]

prod_name = []
price = []
rating = []
del_result = []
for i in final_pg:
    driver.get(i)
    try:

        pinbar = driver.find_element_by_xpath(
            "//div[@class='col col-12-12']//form[@class='EJrIpC']/input[@id='pincodeInputId']")
        pinbar.send_keys('571301')
        pinbar.submit()
        sleep(1)
        prod_name.append(
            driver.find_element_by_xpath("//div[@class='_1HmYoV _35HD7C col-8-12']//h1[@class='_9E25nV']//span").text)
        price.append(driver.find_element_by_xpath(
            "//div[@class='_1HmYoV _35HD7C col-8-12']/div[@class='bhgxx2 col-12-12']//div[@class='_1uv9Cb']/div").text)
        rating.append(driver.find_element_by_xpath("//div[@class='niH0FQ _2nc08B']/span[@class='_2_KrJI']/div").text)
        del_result.append(driver.find_element_by_xpath(
            "//div[@id='container']//div[@class='col col-12-12']//div[@class='_29Zp1s']/span").text)
        # print(prod_name,' ',price,' ',rating,' ',del_result)
    except NoSuchElementException:
        pass

zippedList = list(zip(prod_name, price, rating, del_result))
df = pd.DataFrame(zippedList, columns=['Name', 'Price', 'Rating', 'Delivery date'])
driver.close()

df['Price'] = df['Price'].str.replace('₹','').str.replace(',', '').astype(int)
df.to_excel(r'C:\Users\User\PycharmProjects\Web Scraping\Output.xlsx', index = None, header=True)

wb= openpyxl.load_workbook(r'C:\Users\User\PycharmProjects\Web Scraping\Output.xlsx')
sheet = wb.active
chart = BarChart()
chart.type = "col"
chart.style = 15
chart.title = "Bar Chart"
chart.y_axis.title = 'Price'
chart.x_axis.title = 'Product name'
categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
data = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)
chart.add_data(data)
chart.set_categories(categories)
chart.shape = 8
sheet.add_chart(chart, "F2")
wb.save(r'C:\Users\User\PycharmProjects\Web Scraping\Output.xlsx')